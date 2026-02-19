#!/usr/bin/env python3
"""
Model Testing Runner — Sliding Window Session Tester

Tests the gap+tab_streak ensemble model against real data.

How it works:
  - Loads a data file (one number per line)
  - Session 1: starts at number 1, plays until +$100 profit or runs out of data
  - Session 2: starts at number 2, same thing
  - Session N: starts at number N, same thing
  - Stops when starting position doesn't have enough numbers left

Each session uses:
  - 14 numbers per spin (TOP_PREDICTIONS_COUNT)
  - $2/number starting bet
  - +$1 after 2 consecutive losses
  - -$1 after 2 consecutive wins
  - $1 minimum bet
  - Session target: +$100
  - No stop loss

Output: Excel file with each session in a separate tab, spin-by-spin detail.

Usage:
  python model_testing/test_runner.py model_testing/test_data/yourfile.txt
"""

import sys
import os
import glob
import datetime
import numpy as np

# Add project root to path
PROJECT_ROOT = os.path.join(os.path.dirname(__file__), '..')
sys.path.insert(0, PROJECT_ROOT)

from config import (
    TOTAL_NUMBERS, TOP_PREDICTIONS_COUNT, TOP_PREDICTIONS_MAX,
    BASE_BET, BET_INCREMENT, BET_DECREMENT, MIN_BET, MAX_BET_MULTIPLIER,
    LOSSES_BEFORE_INCREMENT, WINS_BEFORE_DECREMENT,
    SESSION_TARGET, INITIAL_BANKROLL,
    WHEEL_ORDER, WHEEL_TABLE_0, WHEEL_TABLE_19,
)

from app.ml.ensemble import EnsemblePredictor

# Minimum spins needed to make a meaningful prediction
MIN_HISTORY_FOR_PREDICTION = 50  # gap needs 50, tab_streak needs 100 but works with less


def load_data_file(filepath):
    """Load numbers from a data file (one number per line)."""
    numbers = []
    with open(filepath) as f:
        for line in f:
            line = line.strip()
            if line and line.isdigit():
                n = int(line)
                if 0 <= n <= 36:
                    numbers.append(n)
    return numbers


def run_single_session(all_numbers, start_idx, session_id):
    """Run a single session starting from start_idx.

    The model uses numbers from index 0 to start_idx-1 as history,
    then predicts starting from start_idx onward.

    Returns:
        dict with session results and spin-by-spin details
    """
    # Need enough history for the model + enough remaining numbers to play
    if start_idx < MIN_HISTORY_FOR_PREDICTION:
        # Use whatever history is available, but model may be less accurate
        history = all_numbers[:start_idx] if start_idx > 0 else []
    else:
        history = all_numbers[:start_idx]

    remaining = all_numbers[start_idx:]

    if len(remaining) < 5:  # Need at least 5 spins to play a session
        return None

    # Initialize model with history
    predictor = EnsemblePredictor()
    if history:
        predictor.load_history(history)

    # Session state
    bankroll = INITIAL_BANKROLL
    session_profit = 0.0
    bet_per_number = BASE_BET
    consecutive_losses = 0
    consecutive_wins = 0
    spin_details = []
    total_wins = 0
    total_losses = 0

    for spin_idx, actual_number in enumerate(remaining):
        spin_num = spin_idx + 1

        # Get prediction from ensemble
        prediction = predictor.predict()
        predicted_numbers = prediction.get('top_numbers', [])[:TOP_PREDICTIONS_COUNT]
        confidence = prediction.get('confidence', 0)

        # Calculate bet
        num_count = len(predicted_numbers)
        total_bet = bet_per_number * num_count

        # Check if we hit
        hit = actual_number in predicted_numbers

        if hit:
            # Win: payout is 35:1 on the winning number, minus total bet
            winnings = (35 * bet_per_number) - total_bet + bet_per_number  # 35:1 + original bet back
            net = winnings
            session_profit += net
            bankroll += net
            total_wins += 1

            # Consecutive tracking
            consecutive_wins += 1
            consecutive_losses = 0

            # Decrease bet after N consecutive wins
            if consecutive_wins >= WINS_BEFORE_DECREMENT:
                bet_per_number = max(MIN_BET, bet_per_number - BET_DECREMENT)
                consecutive_wins = 0  # Reset counter

            result_str = 'HIT'
        else:
            # Loss: lose total bet
            net = -total_bet
            session_profit += net
            bankroll += net
            total_losses += 1

            # Consecutive tracking
            consecutive_losses += 1
            consecutive_wins = 0

            # Increase bet after N consecutive losses
            if consecutive_losses >= LOSSES_BEFORE_INCREMENT:
                max_bet = BASE_BET * MAX_BET_MULTIPLIER
                bet_per_number = min(max_bet, bet_per_number + BET_INCREMENT)
                consecutive_losses = 0  # Reset counter

            result_str = 'MISS'

        # Record spin detail
        spin_details.append({
            'spin_num': spin_num,
            'actual_number': actual_number,
            'predicted_numbers': ','.join(map(str, sorted(predicted_numbers))),
            'predicted_count': num_count,
            'result': result_str,
            'bet_per_number': bet_per_number if not hit else (bet_per_number if result_str == 'MISS' else spin_details[-1]['bet_per_number'] if spin_details else BASE_BET),
            'total_bet': total_bet,
            'net_profit': net,
            'session_profit': session_profit,
            'bankroll': bankroll,
            'confidence': round(confidence, 1),
            'consec_wins': consecutive_wins,
            'consec_losses': consecutive_losses,
        })

        # Fix bet_per_number in the record (capture BEFORE adjustment)
        spin_details[-1]['bet_per_number'] = round(total_bet / num_count, 2) if num_count > 0 else 0

        # Update model with actual result
        predictor.update(actual_number)

        # Check if target reached
        if session_profit >= SESSION_TARGET:
            break

        # Safety: if bankroll goes to zero, stop
        if bankroll <= 0:
            break

    # Session summary
    session_result = {
        'session_id': session_id,
        'start_index': start_idx,
        'spins_played': len(spin_details),
        'total_wins': total_wins,
        'total_losses': total_losses,
        'win_rate': round(total_wins / len(spin_details) * 100, 1) if spin_details else 0,
        'final_profit': round(session_profit, 2),
        'target_reached': session_profit >= SESSION_TARGET,
        'final_bankroll': round(bankroll, 2),
        'details': spin_details,
    }

    return session_result


def run_all_sessions(filepath):
    """Run sliding window sessions across entire data file."""
    all_numbers = load_data_file(filepath)
    total_numbers = len(all_numbers)
    filename = os.path.basename(filepath)

    print(f"\n{'='*80}")
    print(f"MODEL TEST RUNNER — Sliding Window Session Tester")
    print(f"{'='*80}")
    print(f"Data file: {filename}")
    print(f"Total numbers: {total_numbers}")
    print(f"Model: gap(35%) + tab_streak(22%) + frequency(20%) + wheel(13%) + hot(5%) + pattern(5%)")
    print(f"Settings: {TOP_PREDICTIONS_COUNT} numbers, ${BASE_BET} start, "
          f"+${BET_INCREMENT} after {LOSSES_BEFORE_INCREMENT} losses, "
          f"-${BET_DECREMENT} after {WINS_BEFORE_DECREMENT} wins")
    print(f"Target: +${SESSION_TARGET} per session, No stop loss")
    print(f"{'='*80}\n")

    all_sessions = []
    targets_reached = 0
    total_sessions = 0

    # Minimum 30 spins remaining to start a session
    min_remaining = 30

    for start_idx in range(total_numbers):
        remaining = total_numbers - start_idx
        if remaining < min_remaining:
            print(f"\nStopping: only {remaining} numbers left from index {start_idx}")
            break

        session_id = start_idx + 1
        total_sessions += 1

        # Progress output
        if session_id % 50 == 1 or session_id <= 5:
            print(f"Session {session_id}/{total_numbers - min_remaining}: "
                  f"starting at index {start_idx} "
                  f"(history={start_idx}, remaining={remaining})...",
                  end='', flush=True)

        result = run_single_session(all_numbers, start_idx, session_id)

        if result is None:
            break

        all_sessions.append(result)

        if result['target_reached']:
            targets_reached += 1

        if session_id % 50 == 1 or session_id <= 5:
            status = '✅ +$100' if result['target_reached'] else f'❌ ${result["final_profit"]:.0f}'
            print(f" → {status} in {result['spins_played']} spins "
                  f"({result['win_rate']}% hit rate)")

        # Progress every 100 sessions
        if session_id % 100 == 0:
            pct = targets_reached / total_sessions * 100
            print(f"  ... {total_sessions} sessions done, "
                  f"{targets_reached} hit target ({pct:.1f}%)")

    # Final summary
    print(f"\n{'='*80}")
    print(f"FINAL RESULTS")
    print(f"{'='*80}")
    print(f"Total sessions: {total_sessions}")
    print(f"Target reached (+${SESSION_TARGET}): {targets_reached} ({targets_reached/total_sessions*100:.1f}%)")
    print(f"Target missed: {total_sessions - targets_reached}")

    if all_sessions:
        profits = [s['final_profit'] for s in all_sessions]
        spins = [s['spins_played'] for s in all_sessions]
        win_rates = [s['win_rate'] for s in all_sessions]

        print(f"\nProfit Stats:")
        print(f"  Average: ${np.mean(profits):.2f}")
        print(f"  Median: ${np.median(profits):.2f}")
        print(f"  Best: ${max(profits):.2f}")
        print(f"  Worst: ${min(profits):.2f}")
        print(f"\nSpins per session:")
        print(f"  Average: {np.mean(spins):.1f}")
        print(f"  Min: {min(spins)}")
        print(f"  Max: {max(spins)}")
        print(f"\nHit rate:")
        print(f"  Average: {np.mean(win_rates):.1f}%")
        print(f"  Best: {max(win_rates):.1f}%")
        print(f"  Worst: {min(win_rates):.1f}%")

    return all_sessions, filename


def export_to_csv(all_sessions, filename):
    """Export results to CSV files — one file per session + summary."""
    results_dir = os.path.join(os.path.dirname(__file__), 'results')
    os.makedirs(results_dir, exist_ok=True)

    timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    base_name = os.path.splitext(filename)[0]

    # Try to use openpyxl for Excel with tabs, fall back to CSV
    try:
        import openpyxl
        return export_to_excel(all_sessions, filename, results_dir, timestamp, base_name)
    except ImportError:
        print("\n⚠️  openpyxl not installed. Exporting as CSV files instead.")
        print("  Install with: pip install openpyxl")
        return export_to_csv_files(all_sessions, filename, results_dir, timestamp, base_name)


def export_to_excel(all_sessions, filename, results_dir, timestamp, base_name):
    """Export to Excel with each session in a separate tab."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    output_path = os.path.join(results_dir, f'test_{base_name}_{timestamp}.xlsx')
    wb = openpyxl.Workbook()

    # ─── Styling ───
    header_font = Font(bold=True, size=12)
    hit_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Green
    miss_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # Red
    target_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')  # Yellow
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')  # Blue
    header_text = Font(bold=True, color='FFFFFF', size=11)
    title_font = Font(bold=True, size=14)
    label_font = Font(bold=True, size=11)
    good_font = Font(bold=True, color='008000', size=12)
    bad_font = Font(bold=True, color='CC0000', size=12)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # ─── Overall Stats Sheet (FIRST TAB) ───
    ws_overall = wb.active
    ws_overall.title = 'Overall Stats'

    # Calculate overall statistics
    total_sessions = len(all_sessions)
    sessions_hit = sum(1 for s in all_sessions if s['target_reached'])
    sessions_missed = total_sessions - sessions_hit
    hit_pct = sessions_hit / total_sessions * 100 if total_sessions else 0
    profits = [s['final_profit'] for s in all_sessions]
    spins_list = [s['spins_played'] for s in all_sessions]
    win_rates = [s['win_rate'] for s in all_sessions]
    wins_list = [s['total_wins'] for s in all_sessions]
    losses_list = [s['total_losses'] for s in all_sessions]

    # Title
    ws_overall.cell(row=1, column=1, value='MODEL TEST — OVERALL RESULTS').font = Font(bold=True, size=16)
    ws_overall.cell(row=2, column=1, value=f'Data: {filename}').font = Font(size=11, italic=True)
    ws_overall.cell(row=3, column=1,
                    value=f'Model: gap(35%) + tab_streak(22%) + frequency(20%) + wheel(13%) + hot(5%) + pattern(5%)'
                    ).font = Font(size=10, italic=True)

    # Session Results
    row = 5
    ws_overall.cell(row=row, column=1, value='SESSION RESULTS').font = title_font
    row += 1

    stats = [
        ('Total Sessions Tested', total_sessions),
        ('Sessions HIT +$100 Target', sessions_hit),
        ('Sessions MISSED Target', sessions_missed),
        ('Target Hit Rate', f'{hit_pct:.1f}%'),
        ('', ''),
        ('PROFIT STATS', ''),
        ('Average Profit per Session', f'${np.mean(profits):.2f}'),
        ('Median Profit per Session', f'${np.median(profits):.2f}'),
        ('Best Session Profit', f'${max(profits):.2f}'),
        ('Worst Session Profit', f'${min(profits):.2f}'),
        ('Total Profit (all sessions)', f'${sum(profits):.2f}'),
        ('', ''),
        ('SESSION LENGTH (Spins)', ''),
        ('Average Spins to Hit Target', f'{np.mean(spins_list):.1f}'),
        ('Shortest Session', f'{min(spins_list)} spins'),
        ('Longest Session', f'{max(spins_list)} spins'),
        ('Median Session Length', f'{np.median(spins_list):.0f} spins'),
        ('', ''),
        ('HIT RATE (per spin)', ''),
        ('Average Win Rate', f'{np.mean(win_rates):.1f}%'),
        ('Best Session Win Rate', f'{max(win_rates):.1f}%'),
        ('Worst Session Win Rate', f'{min(win_rates):.1f}%'),
        ('Average Wins per Session', f'{np.mean(wins_list):.1f}'),
        ('Average Losses per Session', f'{np.mean(losses_list):.1f}'),
        ('', ''),
        ('BET SETTINGS', ''),
        ('Numbers per Spin', TOP_PREDICTIONS_COUNT),
        ('Starting Bet/Number', f'${BASE_BET}'),
        ('Bet Increment after Losses', f'+${BET_INCREMENT} after {LOSSES_BEFORE_INCREMENT} consecutive losses'),
        ('Bet Decrement after Wins', f'-${BET_DECREMENT} after {WINS_BEFORE_DECREMENT} consecutive wins'),
        ('Minimum Bet/Number', f'${MIN_BET}'),
        ('Maximum Bet/Number', f'${BASE_BET * MAX_BET_MULTIPLIER}'),
        ('Session Target', f'+${SESSION_TARGET}'),
        ('Starting Bankroll', f'${INITIAL_BANKROLL}'),
    ]

    for label, value in stats:
        if label == '' and value == '':
            row += 1
            continue
        if value == '':
            # Section header
            ws_overall.cell(row=row, column=1, value=label).font = title_font
            row += 1
            continue

        ws_overall.cell(row=row, column=1, value=label).font = label_font
        cell = ws_overall.cell(row=row, column=3, value=value)

        # Color code key metrics
        if 'HIT' in str(label) and 'Target' in str(label):
            cell.font = good_font
        elif 'MISSED' in str(label):
            cell.font = bad_font
        elif 'Hit Rate' in str(label) and '%' in str(value):
            cell.font = good_font if float(value.replace('%', '')) > 50 else bad_font

        row += 1

    # Profit distribution
    row += 2
    ws_overall.cell(row=row, column=1, value='PROFIT DISTRIBUTION').font = title_font
    row += 1

    profit_ranges = [
        ('+$100 or more (target hit)', sum(1 for p in profits if p >= 100)),
        ('+$50 to +$99', sum(1 for p in profits if 50 <= p < 100)),
        ('+$1 to +$49', sum(1 for p in profits if 0 < p < 50)),
        ('$0 (breakeven)', sum(1 for p in profits if p == 0)),
        ('-$1 to -$99', sum(1 for p in profits if -100 < p < 0)),
        ('-$100 to -$299', sum(1 for p in profits if -300 < p <= -100)),
        ('-$300 to -$499', sum(1 for p in profits if -500 < p <= -300)),
        ('-$500 or worse', sum(1 for p in profits if p <= -500)),
    ]

    for col, header in enumerate(['Range', 'Count', '% of Sessions'], 1):
        cell = ws_overall.cell(row=row, column=col, value=header)
        cell.font = header_text
        cell.fill = header_fill
        cell.border = thin_border
    row += 1

    for label, count in profit_ranges:
        ws_overall.cell(row=row, column=1, value=label).border = thin_border
        ws_overall.cell(row=row, column=2, value=count).border = thin_border
        pct = f'{count / total_sessions * 100:.1f}%' if total_sessions else '0%'
        ws_overall.cell(row=row, column=3, value=pct).border = thin_border

        if count > 0 and '+$100' in label:
            for c in range(1, 4):
                ws_overall.cell(row=row, column=c).fill = hit_fill
        elif count > 0 and '-$' in label:
            for c in range(1, 4):
                ws_overall.cell(row=row, column=c).fill = miss_fill
        row += 1

    # Column widths
    ws_overall.column_dimensions['A'].width = 40
    ws_overall.column_dimensions['B'].width = 5
    ws_overall.column_dimensions['C'].width = 30

    # ─── Summary Sheet ───
    ws_summary = wb.create_sheet(title='All Sessions')

    # Summary headers
    summary_headers = [
        'Session', 'Start Index', 'Spins Played', 'Wins', 'Losses',
        'Win Rate %', 'Final Profit $', 'Target Reached', 'Final Bankroll $'
    ]
    for col, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.font = header_text
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # Summary data
    for row_idx, session in enumerate(all_sessions, 2):
        values = [
            session['session_id'],
            session['start_index'],
            session['spins_played'],
            session['total_wins'],
            session['total_losses'],
            session['win_rate'],
            session['final_profit'],
            'YES' if session['target_reached'] else 'NO',
            session['final_bankroll'],
        ]
        for col, val in enumerate(values, 1):
            cell = ws_summary.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

        # Color code: green if target reached, red if not
        if session['target_reached']:
            for col in range(1, len(values) + 1):
                ws_summary.cell(row=row_idx, column=col).fill = hit_fill
        elif session['final_profit'] < 0:
            for col in range(1, len(values) + 1):
                ws_summary.cell(row=row_idx, column=col).fill = miss_fill

    # Auto-width columns
    for col in range(1, len(summary_headers) + 1):
        ws_summary.column_dimensions[
            openpyxl.utils.get_column_letter(col)
        ].width = 16

    # ─── Individual Session Sheets ───
    # Limit to first 250 sessions (Excel has sheet name limits)
    max_sheets = min(len(all_sessions), 250)

    for i, session in enumerate(all_sessions[:max_sheets]):
        sheet_name = f"S{session['session_id']}"
        # Excel sheet names max 31 chars
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]

        ws = wb.create_sheet(title=sheet_name)

        # Session info header
        ws.cell(row=1, column=1, value=f"Session {session['session_id']}").font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value=f"Start Index: {session['start_index']}")
        ws.cell(row=2, column=3, value=f"Result: {'TARGET REACHED ✅' if session['target_reached'] else 'NOT REACHED ❌'}")
        ws.cell(row=3, column=1, value=f"Final Profit: ${session['final_profit']:.2f}")
        ws.cell(row=3, column=3, value=f"Win Rate: {session['win_rate']}%")

        # Detail headers
        detail_headers = [
            'Spin #', 'Actual Number', 'Result', 'Bet/Number $',
            'Total Bet $', 'Net Profit $', 'Session Profit $',
            'Bankroll $', 'Confidence %', 'Predicted Numbers',
            'Consec Wins', 'Consec Losses'
        ]
        header_row = 5
        for col, header in enumerate(detail_headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = header_text
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

        # Detail data
        for detail_idx, detail in enumerate(session['details']):
            row = header_row + 1 + detail_idx
            values = [
                detail['spin_num'],
                detail['actual_number'],
                detail['result'],
                detail['bet_per_number'],
                detail['total_bet'],
                detail['net_profit'],
                round(detail['session_profit'], 2),
                round(detail['bankroll'], 2),
                detail['confidence'],
                detail['predicted_numbers'],
                detail['consec_wins'],
                detail['consec_losses'],
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')

            # Color: green for HIT, red for MISS
            if detail['result'] == 'HIT':
                for col in range(1, len(values) + 1):
                    ws.cell(row=row, column=col).fill = hit_fill
            else:
                for col in range(1, len(values) + 1):
                    ws.cell(row=row, column=col).fill = miss_fill

        # Auto-width
        for col in range(1, len(detail_headers) + 1):
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col)
            ].width = 18

        # Wider for predicted numbers column
        ws.column_dimensions['J'].width = 45

    wb.save(output_path)
    print(f"\n📊 Excel report saved: {output_path}")
    print(f"   Summary tab + {max_sheets} session tabs")
    if len(all_sessions) > max_sheets:
        print(f"   ⚠️  Only first {max_sheets} of {len(all_sessions)} sessions included (Excel limit)")
    return output_path


def export_to_csv_files(all_sessions, filename, results_dir, timestamp, base_name):
    """Fallback: export as individual CSV files."""
    import csv

    # Summary CSV
    summary_path = os.path.join(results_dir, f'test_{base_name}_{timestamp}_summary.csv')
    with open(summary_path, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow([
            'Session', 'Start Index', 'Spins Played', 'Wins', 'Losses',
            'Win Rate %', 'Final Profit $', 'Target Reached', 'Final Bankroll $'
        ])
        for session in all_sessions:
            writer.writerow([
                session['session_id'], session['start_index'],
                session['spins_played'], session['total_wins'],
                session['total_losses'], session['win_rate'],
                session['final_profit'],
                'YES' if session['target_reached'] else 'NO',
                session['final_bankroll'],
            ])

    # Individual session CSVs (first 100 only)
    max_files = min(len(all_sessions), 100)
    session_dir = os.path.join(results_dir, f'test_{base_name}_{timestamp}_sessions')
    os.makedirs(session_dir, exist_ok=True)

    for session in all_sessions[:max_files]:
        session_path = os.path.join(session_dir, f'session_{session["session_id"]:04d}.csv')
        with open(session_path, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerow([
                'Spin #', 'Actual Number', 'Result', 'Bet/Number $',
                'Total Bet $', 'Net Profit $', 'Session Profit $',
                'Bankroll $', 'Confidence %', 'Predicted Numbers',
                'Consec Wins', 'Consec Losses'
            ])
            for detail in session['details']:
                writer.writerow([
                    detail['spin_num'], detail['actual_number'],
                    detail['result'], detail['bet_per_number'],
                    detail['total_bet'], detail['net_profit'],
                    round(detail['session_profit'], 2),
                    round(detail['bankroll'], 2),
                    detail['confidence'], detail['predicted_numbers'],
                    detail['consec_wins'], detail['consec_losses'],
                ])

    print(f"\n📊 CSV reports saved:")
    print(f"   Summary: {summary_path}")
    print(f"   Sessions: {session_dir}/ ({max_files} files)")
    return summary_path


def main():
    if len(sys.argv) < 2:
        # If no argument, look for files in test_data/
        test_data_dir = os.path.join(os.path.dirname(__file__), 'test_data')
        files = sorted(glob.glob(os.path.join(test_data_dir, '*.txt')))
        if not files:
            print("Usage: python model_testing/test_runner.py <data_file.txt>")
            print(f"\nOr place .txt files in: {test_data_dir}/")
            sys.exit(1)
        filepath = files[0]
        print(f"Auto-detected: {filepath}")
    else:
        filepath = sys.argv[1]

    if not os.path.exists(filepath):
        print(f"Error: File not found: {filepath}")
        sys.exit(1)

    # Run all sessions
    all_sessions, filename = run_all_sessions(filepath)

    if not all_sessions:
        print("No sessions completed!")
        sys.exit(1)

    # Export results
    output_path = export_to_csv(all_sessions, filename)

    print(f"\n✅ Done! Open the file to review spin-by-spin results.")


if __name__ == '__main__':
    main()
